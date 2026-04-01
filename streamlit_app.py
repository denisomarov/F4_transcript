import streamlit as st
import io
from openpyxl import load_workbook

def process_data(FileName, FileName_Sheet):

    # открываем файл Сметы, Форма 4

    wb = load_workbook(filename=FileName)

    # удаляем все листы кроме выбранного листа
    sheets_list = wb.sheetnames
    sheets_list.remove(FileName_Sheet)

    for sheet_name in sheets_list:
        sheet_ws = wb[sheet_name]
        wb.remove(sheet_ws)

    # обработка Формы 4
    ws = wb[FileName_Sheet]

    # убираем объединение ячеек в файле
    for merged_cell in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_cell))

    # инициализируем переменную подсчета строк
    row_number = 1
    found = False

    # удаляем строки до начала таблицы данных, до значения "Номер по порядку" в столбце A
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:

            if cell.value != "Номер по порядку":
                row_number = row_number + 1
            else:
                found = True

        if found: break

    ws.delete_rows(1, row_number - 1)

    # удаляем строки, первые ячейки которых содержат пустые значения и точку (признак промежуточного значения)
    for row in range(ws.max_row, 0, -1):
        if ws.cell(row=row, column=1).value == '':
            ws.delete_rows(row)
        if ws.cell(row=row, column=1).value == None:
            ws.delete_rows(row)
        if '.' in str(ws.cell(row=row, column=1).value):
            ws.delete_rows(row)

    # форматируем таблицу
    for row in range(1, ws.max_row):
        ws.row_dimensions[row].height = 65
        cell_value = str(ws.cell(row=row, column=2).value)
        ws.cell(row=row, column=2).value = cell_value.split(' ')[0]

    # запись результата в буфер
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# --- Интерфейс Streamlit ---
st.title("Выделение оборудования, материалов и работ из Сметы")

uploaded_file = st.file_uploader("Загрузите файл Сметы в формате Эксель", type=['xlsx'])

if uploaded_file is not None:
    wb_temp = load_workbook(uploaded_file, read_only=True)
    sheets = wb_temp.sheetnames
    selected_sheet = st.selectbox("Выберите лист:", sheets)

    if st.button("✨ Обработать файл"):
        # Важно: для openpyxl нужно сбросить указатель файла
        uploaded_file.seek(0)
        result = process_data(uploaded_file, selected_sheet)

        st.download_button(
            label="📥 Скачать готовый файл",
            data=result,
            file_name="Форма 4. Расшифровка.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )